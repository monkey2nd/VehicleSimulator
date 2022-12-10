from pathlib import Path
from typing import Any, List

import numpy as np
import openpyxl as px
from openpyxl.worksheet.worksheet import Worksheet


def make_result(dir_path: Path, penetration_ls, car_max_ls, seed_ls, merging_ls, ctrl_cfgs):
    """
    save.pyは各シミュレーション結果を記録するプログラム
    result.pyはsave.pyにて保存されたファイルから全体の結果を記録するプログラム
    """

    wb_name = "result_" + dir_path.name + ".xlsx"
    data_set = [[[] for _ in range(len(penetration_ls))] for __ in range(len(car_max_ls))]
    wb = px.Workbook()

    col_title_ls = ["[-10]", "[10-20]", "[20-30]", "[30-40]", "40-", "平均占有率"]
    it_max = len(penetration_ls) * len(car_max_ls) * len(ctrl_cfgs) * len(merging_ls)
    it_count = 1

    for ctrl_cfg in ctrl_cfgs:
        for num_m in merging_ls:
            if ctrl_cfg["lc_control"]:
                ws = wb.create_sheet("lc_あり_合流者数" + str(num_m))
            else:
                ws = wb.create_sheet("lc_無し_合流者数" + str(num_m))
            for num_p, penetration in enumerate(penetration_ls):
                col_s = num_p * (len(col_title_ls) + 3) + 1
                for num_c, car_max in enumerate(car_max_ls):
                    use_lc_filename = ""
                    print(str(it_count) + "/" + str(it_max) + "データ形成中")
                    row_s = num_c * (len(seed_ls) + 4) + 1
                    f_path = dir_path / ("普及率" + str(penetration * 100) + "%") / (
                            "車両数" + str(car_max) + "_" + str(num_m))
                    if not penetration == 0:
                        if ctrl_cfg["lc_control"]:
                            f_path /= "lc_controlあり"
                            use_lc_filename = "use_lc"
                        else:
                            f_path /= "lc_control無し"

                    table = get_table(col_title_ls=col_title_ls, seed_ls=seed_ls, car_max=car_max,
                                      penetration=penetration, path=f_path, data_set=data_set, num_p=num_p, num_c=num_c,
                                      use_lc_filename=use_lc_filename)
                    write_list(ws=ws, input_data=table, column=col_s, row=row_s)
                    it_count += 1

            get_table2(ws=ws, data_set=data_set, penetration_ls=penetration_ls, car_max_ls=car_max_ls,
                       col_title_ls=col_title_ls)

    print(str(dir_path) + "/" + wb_name + " を保存中...")
    wb.save(str(dir_path) + "/" + wb_name)
    wb.close()


def write_list(ws: Worksheet, input_data, column, row):
    """
    一次元又は二次元のリストをエクセルに直接書き込める関数
    """
    for index, data in enumerate(input_data):
        if type(data) in (List, list, np.ndarray):
            write_list(ws=ws, input_data=data, column=column, row=row + index)
        else:
            ws.cell(column=column + index, row=row).value = data


def get_template(col_title_ls, seed_ls, car_max, penetration) -> List[List[Any]]:
    """
    記録するデータのテンプレート
    """
    app_list = [None for _ in range(len(col_title_ls))]
    col_title = [str(car_max) + "普及率:" + str(penetration)] + col_title_ls
    template = [col_title]
    for seed in seed_ls:
        template.append([seed] + app_list)
    template.append(["平均値"] + app_list)
    template.append(["標準偏差"] + app_list)

    return template


def get_data(seed_ls, penetration, path: Path, data_set: List[List],
             num_c, num_p, use_lc_filename) -> List[List]:
    """
    減速量に関係するデータを取得する関数
    """
    data = []
    for seed in seed_ls:
        data_tmp = []
        f_path = path / ("seed" + str(seed) + "_penetration" + str(int(penetration * 100)) + use_lc_filename + ".xlsx")
        wb = px.load_workbook(str(f_path), read_only=True, data_only=True)
        for data_sources in wb["減速量"]["G3":"K3"]:
            for data_source in data_sources:
                data_tmp.append(data_source.value)
            if penetration != 0:
                data_tmp.append(wb["車線普及率"]["I2"].value)
        data.append(data_tmp)
        wb.close()

    data_array = np.array(data)
    data_mean = np.mean(data_array, axis=0)
    data_std = np.std(data_array, axis=0)
    data_std *= 2  # ? 95%信頼区間
    data_array = np.append(data_array, [data_mean], axis=0)
    data_array = np.append(data_array, [data_std], axis=0)
    data_set[num_c][num_p] = data_mean.tolist() + data_std.tolist()
    return data_array.tolist()


def get_table(col_title_ls, seed_ls, car_max, penetration, data_set, num_c, num_p, path, use_lc_filename):
    template = get_template(col_title_ls=col_title_ls, seed_ls=seed_ls, car_max=car_max, penetration=penetration)
    datas = get_data(path=path, seed_ls=seed_ls, penetration=penetration, data_set=data_set, num_p=num_p, num_c=num_c,
                     use_lc_filename=use_lc_filename)
    for row_num, data_ls in enumerate(datas):
        for col_num, value in enumerate(data_ls):
            template[row_num + 1][col_num + 1] = value

    return template


def get_template2(col_title_ls, car_max, penetration_ls):
    app_list = [None for _ in range(len(col_title_ls))]
    col_title = [car_max] + col_title_ls + ["分散", None, None, None]
    template = [col_title]
    for penetration in penetration_ls:
        template.append(["普及率" + str(int(penetration * 100))] + app_list + app_list)

    return template


def get_table2(ws: Worksheet, data_set, penetration_ls, car_max_ls, col_title_ls):
    for car_max_index in range(len(car_max_ls)):
        template = get_template2(col_title_ls=col_title_ls, car_max=car_max_ls[car_max_index],
                                 penetration_ls=penetration_ls)

        for row_num, data_ls in enumerate(data_set[car_max_index]):
            for col_num, value in enumerate(data_ls):
                template[row_num + 1][col_num + 1] = value

        write_list(ws=ws, input_data=template, column=(len(col_title_ls) + 3) * len(penetration_ls) + 1,
                   row=((len(penetration_ls) + 3) * car_max_index) + 1)
