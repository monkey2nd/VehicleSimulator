from pathlib import Path
from typing import Dict, List

import numpy as np
import openpyxl as px
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from Class_dir.Controller import Controller
from Class_dir.DataCollect import DataCollect
from Class_dir.LaneManager import LaneManager
from Class_dir.VehicleClass import Vehicle
from Class_dir.VehlogClass import Vehlog, Vehtpl


def search_veh(car: List[List[Vehicle]], time, veh_id) -> Vehicle:
    return next((check_car for check_car in car[time] if check_car.veh_id == veh_id), None)


def write_list(ws: Worksheet, written_data, column=1, row=1):
    """
    一次元又は二次元のリストをエクセルに直接書き込める関数
    返却値は次のrow(行）
    """
    for index, data in enumerate(written_data):
        if type(data) in (List, list, np.ndarray):
            write_list(ws=ws, written_data=data, column=column, row=row + index)
        else:
            ws.cell(column=column + index, row=row).value = data
    return row + 1


def create_path(controller: Controller, lm: LaneManager, seed, dir_path: Path):
    """
    保存先ファイル名を生成する関数
    """
    use_lc_filename = ""
    f_path = dir_path / ("普及率" + (str(lm.penetration * 100) + "%")) / (
            "車両数" + str(lm.q_main_lane) + "_" + str(lm.q_lane_ls[0]))

    if controller.use_control:
        if controller.use_lc_control is True:
            f_path /= "lc_controlあり"
            use_lc_filename = "use_lc"
        else:
            f_path /= "lc_control無し"

    f_path /= ("seed" + str(seed) + "_penetration" + str(int(lm.penetration * 100)) + use_lc_filename + ".xlsx")

    return f_path


def create_excel_file():
    wb = px.Workbook()
    ws = wb.active
    return wb, ws


def create_info_sheet(wb: Workbook, lm: LaneManager, time_max):
    ws = wb.create_sheet("情報")
    data = [
        ["シミュレーション時間", time_max / 10],
        ["加速交通量", lm.get_q(0)],
        ["第一走行交通量", lm.get_q(1)],
        ["第二走行交通量", lm.get_q(2)],
        ["追い越し交通量", lm.get_q(3)],
        ["加速車線開始[m]", lm.ms_start],
        ["加速車線終了[m]", lm.ms_end],
        ["基地局", lm.controller.use_control],
        ["合流車両の割合", lm.merging_ratio]
        ]
    write_list(ws=ws, written_data=data)


def get_veh_lane(vehlog: Vehlog, lane):
    id_list = []

    for veh_time in vehlog.log:
        for vehicle in veh_time.values():
            if vehicle.lane == lane:
                if vehicle.veh_id not in id_list:
                    id_list.append(vehicle.veh_id)

    return id_list


def save4_write0(ws: Worksheet, vehlog: Vehlog, time_max, lane, lane_id_list, acceleration_lane_start,
                 acceleration_lane_end):  # save4専用の関数
    sum1 = 0
    sum2 = 0
    t1 = 0
    t2 = 0
    avg_row = 2
    ws.freeze_panes = 'A2'  # 先頭行固定
    if 0 in lane_id_list:  # Id=0が含まれていた場合
        lane_id_list.remove(0)
    row_tmp = 1

    for column, veh_id in enumerate(lane_id_list, 1):
        ws.cell(row=row_tmp, column=column).value = veh_id  # 一番上の行にIDを記録する
    row_tmp += 1

    for time in range(time_max):
        for column, veh_id in enumerate(lane_id_list, 1):
            check_veh = vehlog.get(time, veh_id)
            if check_veh is not None:
                if check_veh.lane == lane:
                    ws.cell(row=row_tmp, column=column).value = check_veh.vel_h
                    if acceleration_lane_start - 300 < check_veh.front < acceleration_lane_end:
                        sum1 += check_veh.vel_h
                        t1 += 1
                    if acceleration_lane_start < check_veh.front < acceleration_lane_end:  # ? 合流部
                        sum2 += check_veh.vel_h
                        t2 += 1
                else:
                    ws.cell(row=row_tmp + 1, column=column).value = ""
        row_tmp += 1

    ws.cell(row=avg_row, column=1).value = "平均速度"
    ws.cell(row=avg_row, column=2).value = (sum1 / t1)
    ws.cell(row=avg_row, column=3).value = (sum2 / t2)


def create_lane_vel_sheet(wb: Workbook, vehlog: Vehlog, lm: LaneManager, time_max):
    lane0_id_list = get_veh_lane(vehlog, 0)  # 加速車線を走行した車両のIDを調べる
    lane1_id_list = get_veh_lane(vehlog, 1)  # 第一走行車線を走行した車両のIDを調べる
    lane2_id_list = get_veh_lane(vehlog, 2)  # 第二走行車線を走行した車両のIDを調べる
    lane3_id_list = get_veh_lane(vehlog, 3)  # 追越車線を走行した車両のIDを調べる
    ws = wb.create_sheet(title="加速速度")  # シート名を指定
    save4_write0(ws, vehlog, time_max, 0, lane0_id_list, lm.ms_start, lm.ms_end)
    ws = wb.create_sheet(title="第一走行速度")  # シート名を指定
    save4_write0(ws, vehlog, time_max, 1, lane1_id_list, lm.ms_start, lm.ms_end)
    ws = wb.create_sheet(title="第二走行速度")  # シート名を指定
    save4_write0(ws, vehlog, time_max, 2, lane2_id_list, lm.ms_start, lm.ms_end)
    ws = wb.create_sheet(title="追越速度")  # シート名を指定
    save4_write0(ws, vehlog, time_max, 3, lane3_id_list, lm.ms_start, lm.ms_end)


def create_merging_info_sheet(wb: Workbook, vehlog: Vehlog, dc: DataCollect, lm: LaneManager, time_max):
    # ここから加速車線から走行車線に移動するまでにかかった時間を記録する
    ws = wb.create_sheet(title="合流にかかった時間")
    fill = PatternFill(patternType='solid', fgColor='FF0000')
    title_ls = ["ID", "Type", "車線変更時刻", "車両生成から車線変更までの時間", "位置", "譲った車両ID", "前方通信位置",
                "後方通信位置", "希望速度", "合流時車両速度", "1秒前合流時車両速度", "2秒前合流時車両速度",
                "3秒前合流時車両速度", "4秒前合流時車両速度"]
    row = write_list(ws=ws, written_data=title_ls)
    time_tmp = 0
    sum1 = 0
    sum2 = 0
    sum_tmp = 0
    # ##合流できた車両の合流までにかかった時間を調べる###
    check_veh_old = None
    check_veh = None
    for veh_id in range(1, lm.get_veh_max):  # 車両IDが0の車は除く
        for time in range(time_max):
            check_veh = vehlog.get(time, veh_id)
            if check_veh is not None:
                if lm.ms_start < check_veh.back:
                    if check_veh.lane == 0:
                        time_tmp += 1  # 合流区間を走った時間を記録している
                        check_veh_old = check_veh
                    else:
                        lc_time = time
                        break

        # 加速車線を走っていた時間が0秒より長く、最後の時間の車線が加速車線でないとき、合流できた車両とみなす

        if time_tmp > 0 and check_veh is not None and check_veh.lane != 0:
            data_ls = [veh_id, check_veh.type, lc_time, time_tmp, check_veh.front, check_veh_old.app_car_id,
                       dc.get_cd(id=check_veh.veh_id).cep, dc.get_cd(id=check_veh.veh_id).csp, check_veh.vd_h]
            if lm.ms_end - 50 < check_veh.front:  # 950m以降の車線変更ではセル色を変える
                ws.cell(row=row, column=5).fill = fill
            for k in range(0, 5):
                data_ls.append(vehlog.get(lc_time - k, veh_id).vel_h)  # 4秒前までの合流時車両速度
                if check_veh.vel < 60 / 3.6 and k == 0:  # ! 20km/s以下の車線変更ではセル色を変える changed by koki
                    ws.cell(row=row, column=10).fill = fill
            if lc_time >= 300:
                sum1 += time_tmp
                sum2 += check_veh.front
                sum_tmp += 1
            row = write_list(ws=ws, written_data=data_ls, row=row)
        time_tmp = 0
    if sum_tmp != 0:
        ave_time = sum1 / sum_tmp
        ave_length = sum2 / sum_tmp
        data_ls = [["300秒以降の車両の時間平均値", "300秒以降の車両の距離平均値"],
                   [ave_time / 10, ave_length - lm.ms_start]]
        row = write_list(ws=ws, written_data=data_ls, row=row)
    time_tmp = 0
    sum1 = 0
    sum2 = 0
    sum_tmp = 0
    for car_id in range(1, lm.get_veh_max):  # 車両IDが0の車は除く
        for time in range(time_max):
            check_veh = vehlog.get(time, car_id)
            if check_veh is not None:
                if check_veh.lane == 0:
                    time_tmp += 1  # 加速車線を走った時間を記録している
                elif check_veh.lane == 1 or check_veh.lane == 2:
                    break  # 走行車線を走った場合はそれ以上調べる必要はない
                elif check_veh.lane == 3:
                    break  # 追越車線を走った場合はそれ以上調べる必要はない
        # 加速車線を走っていた時間が0秒より長く、最後の時間の車線が加速車線でないとき、合流できた車両とみなす
        if time_tmp > 0 and check_veh is not None and check_veh.lane != 0:
            if lc_time >= 300:
                sum1 += (time_tmp - ave_time) ** 2
                sum2 += (check_veh.front - ave_length) ** 2
                sum_tmp += 1
        time_tmp = 0
    if sum_tmp != 0:
        data_ls = [["300秒以降の車両の時間平均値の標準偏差", "300秒以降の車両の距離平均値の標準偏差"],
                   [(sum1 / sum_tmp) ** 0.5, (sum2 / sum_tmp) ** 0.5]]
        write_list(ws=ws, written_data=data_ls, row=row)


def colorBarRGB(car_id):  # 可視化するのに使う関数
    global color
    car_id = car_id % 10
    if car_id == 1:
        color = "FF0000"  # 赤
    elif car_id == 2:
        color = "FFA500"  # オレンジ
    elif car_id == 3:
        color = "00FF00"  # 黄緑
    elif car_id == 4:
        color = "007400"  # 緑
    elif car_id == 5:
        color = "00FFF"  # 水色
    elif car_id == 6:
        color = "0000FF"  # 青
    elif car_id == 7:
        color = "8D0093"  # 紫
    elif car_id == 8:
        color = "FF00FF"  # ピンク
    elif car_id == 9:
        color = "800000"  # 茶色
    elif car_id == 0:
        color = "808080"  # グレー

    return color


def colorBarRGB2(type_, shift_lane: bool):  # 可視化するのに使う関数
    if type_ == 0:
        _color = "808080"
    elif type_ == 1:
        _color = "00FFFF"

    if shift_lane:
        _color = "FFA500"

    return _color


def abc_from_number(number):  # 可視化するのに使う関数
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    colname = ""
    divend = number
    while divend > 0:
        modulo = (divend - 1) % 26
        colname = alphabet[modulo] + colname
        divend = (divend - modulo) // 26
    return colname


def create_visual_sheet(wb: Workbook, vehlog: Vehlog, lm: LaneManager, time_max):
    def draw_vehicle(ws: Worksheet, vehicle: Vehtpl, setting_dic: Dict[str, Border], row, column):
        _color = colorBarRGB2(vehicle.type, vehicle.shift_lane)
        ws.cell(row=row, column=column).font = Font(b=True, color=_color)
        ws.cell(row=row, column=column).value = str(vehicle.veh_id) + str("/") + str(int(vehicle.vel_h))  # IDを記録1
        for i in range(5):  # 車両長分ループ
            if i == 0:
                ws.cell(row=row - i, column=column).border = setting_dic["bottom_border"]
            elif i == 4:
                ws.cell(row=row - i, column=column).border = setting_dic["top_border"]
            else:
                ws.cell(row=row - i, column=column).border = setting_dic["middle_border"]

    lane_num = 4  # *　レーン数
    border = Border(right=Side(style='thin', color='000000'))
    ws = wb.create_sheet(title="可視化")
    lane_id_list = [0] * 2000
    fill = PatternFill(patternType='solid', fgColor='FF0000')

    veh_side = Side(style="thin", color="000000")
    setting_dict = {"top_border": Border(top=veh_side, right=veh_side, left=veh_side),
                    "bottom_border": Border(bottom=veh_side, right=veh_side, left=veh_side),
                    "middle_border": Border(right=veh_side, left=veh_side)
                    }

    for time in tqdm(range(time_max), desc='visualize'):
        vehicle_timelog = vehlog.get(time)
        for car_id in range(1, lm.road_length):  # 罫線引くところ 20200615
            ws.cell(row=car_id, column=lane_num + lane_num * time).border = border
        if not vehicle_timelog == {}:
            ws.cell(row=1, column=1 + lane_num * time).value = time  # IDを記録1
            for check_vehicle in vehicle_timelog.values():
                lane = check_vehicle.lane
                draw_vehicle(ws=ws, vehicle=check_vehicle, row=int(check_vehicle.front),
                             setting_dic=setting_dict,
                             column=check_vehicle.lane + 1 + lane_num * time)

                # 車線が変わった時だけセル色を変える
                if lane != lane_id_list[check_vehicle.veh_id] and lane_id_list[check_vehicle.veh_id] != 0:
                    ws.cell(row=int(check_vehicle.front),
                            column=check_vehicle.lane + 1 + lane_num * time).fill = fill
                if check_vehicle.shift_begin_time + 30 < time and lane == 1 and check_vehicle.shift_lane is True:
                    ws.cell(row=int(check_vehicle.front),
                            column=check_vehicle.lane + 1 + lane_num * time).fill = fill
                lane_id_list[check_vehicle.veh_id] = lane

    for time in range(time_max):
        col = abc_from_number(time + 1)
        ws.column_dimensions[col].width = 8
    ws.freeze_panes = 'A2'  # 先頭行固定
    return ws


def create_log_sheet(wb: Workbook, vehlog: Vehlog, time_max):
    ws: Worksheet = wb.create_sheet(title="log_sheet")
    row = 1

    list_ = ["時間", "Id", "前方位置", "車線", "車両タイプ", "加速度", "加速理由", "速度", "希望速度", "車線変更閾値",
             "車間", "希望車間距離", "相対速度", "希望車頭時間", "前方車両ID", "後方車両ID", "目標車両ID",
             "前方車線変更車両", "車線変更途中", "車線変更先", "車線変更開始時間", "車線変更先車間距離", "mode", "ego"]

    for col_tmp, list_tmp in enumerate(list_, 1):  # 1行目の文字の部分の書き込み col_tmpは1からスタートする
        ws.cell(row=row, column=col_tmp).value = list_tmp
    row = 2
    ws.freeze_panes = 'A2'  # 先頭行固定
    for time in tqdm(range(0, time_max), desc="save_log"):
        for vehicle in vehlog.get_logvalues(time):  # ID0は除外,車の台数分ループ
            ws.cell(row=row, column=1).value = (time / 10)  # 最も左の列の時間の部分
            # col = abc_from_number(vehicle.lane + 1 + 4 * time)
            # ws.cell(row=row, column=2).hyperlink = "#可視化!" + str(col) + str(int(vehicle.front))
            data_ls = [time, vehicle.veh_id, vehicle.front, vehicle.lane, vehicle.type, vehicle.accel,
                       vehicle.accel_name, round(vehicle.vel_h, 2), round(vehicle.vd_h, 2), round(vehicle.vdcl_h, 2),
                       vehicle.distance, vehicle.desired_distance, round(vehicle.delta_v, 2), vehicle.tau,
                       vehicle.front_car_id, vehicle.back_car_id, vehicle.target_car_id, vehicle.shift_front_veh_id,
                       vehicle.shift_lane, vehicle.shift_lane_to, vehicle.shift_begin_time,
                       vehicle.shift_distance_go, vehicle.mode, vehicle.ego]

            row = write_list(ws=ws, written_data=data_ls, column=1, row=row)
    ws.auto_filter.ref = get_column_letter(1) + str(1) + ':' + get_column_letter(ws.max_column) + str(ws.max_row)


def time_vel_sheet(wb: Workbook, car: List[List[Vehicle]], car_max):
    time_max = 600 * 10
    ws = wb.create_sheet(title='平均速度推移')
    row = 1
    list_ = ['時間', '合流車線', '第一走行車線', '第二走行車線', '追越車線']

    for column, col_name in enumerate(list_, 1):
        ws.cell(row=row, column=column).value = col_name

    ws.freeze_panes = 'A2'

    for time in range(time_max + 1):  # ? (10秒毎にループ)(入力で調整)
        lane_avgv = [[0, 0] for _ in range(len(list_) - 1)]  # ? [[合計車両数、合計車両速度],,,←車両レーン数]
        car_tmp = car[time]  # ? アクセス数削減の処理
        for car_id in range(1, car_max):
            check_car = car_tmp[car_id]
            if check_car.lane == -1:
                continue
            else:
                lane_avgv[check_car.lane][0] += 1
                lane_avgv[check_car.lane][1] += check_car.vel_h

        for column, value_ls in enumerate(lane_avgv, 2):
            value = value_ls[1] / value_ls[0] if not value_ls[0] == 0 else None
            ws.cell(row=time + 2, column=1).value = time / 10
            ws.cell(row=time + 2, column=column).value = value


def time_avgvel_sheet(wb: Workbook, car: List[List[Vehicle]]):
    time_max = 600 * 10
    ws = wb.create_sheet(title='秒間平均速度')
    column_list = ['合流車線', '第一走行車線', '第二走行車線', '追越車線']
    row_list = ["0秒以降平均", '100秒以降平均', '200秒以降平均', '300秒以降平均', '400秒以降平均', '500秒以降平均']
    lane_avgv = [[0, 0] for _ in range(len(column_list))]  # ? [[合計車両数、合計車両速度],,,←車両レーン数]

    row = 1
    for column, column_title in enumerate(column_list, 2):
        ws.cell(row=row, column=column).value = column_title

    column = 1
    for row, row_title in enumerate(row_list, 2):
        ws.cell(row=row, column=column).value = row_title

    row = len(row_list) + 1

    for time in range(time_max, -1, -10):
        for check_car in car[time]:
            lane_avgv[check_car.lane][0] += 1
            lane_avgv[check_car.lane][1] += check_car.vel

        if time != time_max and time % (100 * 10) == 0:
            for column, values in enumerate(lane_avgv, 2):
                value = values[1] / values[0] if not values[0] == 0 else None
                ws.cell(row=row, column=column).value = value
            row -= 1


def moving_avg_sheet(wb: Workbook, vehlog: Vehlog, time_max):
    ws = wb.create_sheet('移動平均速度')
    column_titles = ['時間', '合流車線', '第一走行車線', '第二走行車線', '追越車線']
    row = 2
    lane_num = len(column_titles) - 1
    window = 10  # ? 移動平均の幅
    moving_avg = [[] for _ in range(lane_num)]  # ? [[[合計車両数、合計車両速度]×Window]×レーン数]

    for column, column_title in enumerate(column_titles, 1):
        ws.cell(row=1, column=column).value = column_title

    ws.freeze_panes = 'A2'

    for time in range(time_max):  # ? 1秒毎でループ
        lane_vel_avg = [[0, 0] for _ in range(lane_num)]  # ? [[合計車両数、合計車両速度],,,←レーン数]
        for check_veh in vehlog.get_logvalues(time):
            if check_veh.lane == -1:
                continue
            else:
                lane_vel_avg[check_veh.lane][0] += 1
                lane_vel_avg[check_veh.lane][1] += check_veh.vel_h

        for lane in range(lane_num):
            moving_avg[lane].append(lane_vel_avg[lane])
            if len(moving_avg[lane]) > window:
                moving_avg[lane].pop(0)

        if time >= window:
            written_data = [time]
            for lane in range(lane_num):
                veh_num = 0
                veh_vel = 0
                for i in range(window):
                    veh_num += moving_avg[lane][i][0]
                    veh_vel += moving_avg[lane][i][1]
                if veh_num == 0:
                    written_data.append(-1)
                else:
                    written_data.append(veh_vel / veh_num)
            row = write_list(ws=ws, written_data=written_data, row=row)


def deceleration_log_sheet(wb: Workbook, dc: DataCollect, lm: LaneManager):
    ws = wb.create_sheet("減速量")
    category = ['Id', '初期速度', '最小速度', '減速量1', '減速量2']
    max_range = 4
    bandwidth = 10
    band_count = [0 for _ in range(max_range + 1)]

    for column, value in enumerate(category, 1):
        ws.cell(row=1, column=column).value = value

    row = 2

    for car_id in range(1, lm.get_veh_max):
        if car_id not in lm.second_control_car_ls:
            ws.cell(row=row, column=1).value = car_id
            ws.cell(row=row, column=2).value = dc.dece_ls[car_id].v_init_h
            ws.cell(row=row, column=3).value = dc.dece_ls[car_id].min_vel_h
            ws.cell(row=row, column=4).value = dc.get_v_diff_h(id=car_id)

            band = int(dc.get_v_diff_h(id=car_id) / bandwidth)
            band = band if not band > max_range else max_range

            band_count[band] += 1

            ws.cell(row=row, column=5).value = band

            row += 1

    row = 1

    for column, value in enumerate(band_count, len(category) + 2):
        if not column == max_range + len(category) + 2:
            ws.cell(row=1, column=column).value = "-" + str((column - len(category) - 2) * bandwidth + bandwidth)
        else:
            ws.cell(row=1, column=column).value = str(max_range * bandwidth) + "-"

        ws.cell(row=2, column=column).value = value
        ws.cell(row=3, column=column).value = value / lm.get_q(0)
    ws.auto_filter.ref = "E1:E" + str(ws.max_row)


def lane_penetration_log(wb: Workbook, vehlog: Vehlog, lm: LaneManager, time_max):
    """
    各車線の合流部手前までの自動運転車両の占有率を保存
    """
    ws = wb.create_sheet("車線普及率")
    category = ["時間(s)", "合流車線", "第一走行車線", "第二走行車線", "追い越し車線"]
    table = [category]
    datas = []

    for time in range(time_max):
        if vehlog.get_len(time):
            data_ls = [0 for _ in range(len(category))]
            data_ls[0] = time
            data_ls_tmp = [0 for _ in range(len(category) - 1)]  # 各車線の自動運転車両数
            data_ls_tmp2 = [0 for _ in range(len(category) - 1)]  # 各車線の合計車両数

            for check_car in vehlog.get_logvalues(time):
                if check_car.front < lm.ms_start:
                    if check_car.type == 1:
                        data_ls_tmp[check_car.lane] += 1
                    data_ls_tmp2[check_car.lane] += 1

            for index_, (auto_num, sum_num) in enumerate(zip(data_ls_tmp, data_ls_tmp2)):
                if not sum_num == 0:
                    data_ls[index_ + 1] = auto_num / sum_num
                else:
                    data_ls[index_ + 1] = -1

            datas.append(data_ls)

    table.append(datas)
    write_list(ws=ws, written_data=table, column=1, row=1)

    # 平均値等を求める
    category2 = [None, "合流", "第一走行", "第二走行", "追い越し"]
    data_arr = np.array(datas)
    data_arr = np.where(data_arr == -1, np.nan, data_arr)  # -1の情報だけ消す

    avg_ls = ["平均値"] + np.nanmean(data_arr[:, 1:], axis=0).tolist()
    std_ls = ["標準偏差"] + np.nanstd(data_arr[:, 1:], axis=0).tolist()
    write_list(ws=ws, written_data=[category2, avg_ls, std_ls], column=len(category) + 2, row=1)


def tracking_log(wb: Workbook, vehlog: Vehlog, id_list: List[int], time_max: int, sheet_title="スペースダイアグラム"):
    ws = wb.create_sheet(sheet_title)
    row = 1
    title = ["time", "id", "front", "vel"]
    row = write_list(ws=ws, written_data=title, column=1, row=row)
    for time in range(0, time_max):
        for veh_id in id_list:
            veh_data = vehlog.get(time, veh_id)
            if veh_data is not None:
                row = write_list(ws=ws, written_data=[time, veh_data.veh_id, veh_data.front, veh_data.vel_h], column=1,
                                 row=row)

    pass


def create_avg_vel_log(wb: Workbook, vehlog: Vehlog, id_ls=None, sheet_title="平均速度分布"):
    ws = wb.create_sheet(sheet_title)
    column_title = ["id", "平均速度"]
    vel_dict: Dict[int, np.ndarray] = {}
    row = 1
    row = write_list(ws=ws, written_data=column_title, row=row, column=1)
    for time_log in vehlog.log:
        for vehicle in time_log.values():
            if id_ls is None or vehicle.veh_id in id_ls:
                if vehicle.veh_id not in vel_dict.keys():
                    vel_dict[vehicle.veh_id] = np.array([vehicle.vel_h])
                else:
                    vel_dict[vehicle.veh_id] = np.append(vel_dict[vehicle.veh_id], vehicle.vel_h)

    for veh_id in vel_dict.keys():
        row = write_list(ws=ws, written_data=[veh_id, vel_dict[veh_id].mean()], row=row, column=1)

    '''
def deceleration_log_sheet(wb: Workbook, ws: Worksheet, vehlog: List[List[Vehicle]], car_max, 
time_max, q_lane0, interval=10):  
# ? 合流車両の後続の車両がどれくらい減速しているか
    time_max = 600 * 10
    ws = wb.create_sheet("減衰グラフ")
    category = ['Id', '合流時速度', '合流後最小速度', '減速量']
    lane_id_list = [-1 for i in range(car_max + 1)]
    back_car_id_list = []

    for column, column_title in enumerate(category, 1):
        ws.cell(row=1, column=column).value = column_title
    row = 2

    for time in range(interval, time_max, interval):
        run_car_ = vehlog[time]

        for car_id in range(car_max):
            vehicle = run_car_[car_id]

            if lane_id_list[vehicle.Id] == -1:
                lane_id_list[vehicle.Id] = vehicle.lane

            elif not lane_id_list[vehicle.Id] == vehicle.lane:
                if lane_id_list[vehicle.Id] == 0 and vehicle.lane == 1 and vehicle.back_car is not None:
                    back_car_id_list.append([vehicle.back_car_id, vehicle.back_car.vel_h, vehicle.back_car.vel_h])  # ? [[合流先後方車両、合流時後方車両速度、合流後最小車両速度]....]
                lane_id_list[vehicle.Id] = vehicle.lane

        for back_car in back_car_id_list:
            back_car_id = back_car[0]
            back_car_vel_h_first = back_car[1]
            back_car_vel_h_min = back_car[2]

            if not vehlog[time][back_car_id].lane == -1:
                if vehlog[time][back_car_id].vel_h < back_car_vel_h_min:
                    back_car_vel_h_min = vehlog[time][back_car_id].vel_h

            back_car[2] = back_car_vel_h_min

    for back_car in back_car_id_list:
        print(back_car)
        for column, back_car_data in enumerate(back_car, 1):
            ws.cell(row=row, column=column).value = back_car_data
        ws.cell(row=row, column=column + 1).value = back_car[1] - back_car[2]
        ws.cell(row=row, column=column + 2).value = int((back_car[1] - back_car[2]) / 5)  # ? 5km/h 間隔でグラフを作成するため
        row += 1

    ws.cell(row=1, column=len(category) + 1).value = q_lane0
    return ws
'''


"""
def save4(car: List[List[Vehicle]], lm: LaneManager, time_max, seed, dir_name) -> None:
    path = create_path(seed=seed, lm=lm, dir_name=dir_name)
    wb, ws = create_excel_file()
    ws.title = "情報"  # 最初は'sheet1'というシートがデフォルトで作成されるため、名前変更
    create_info_sheet(wb=wb, lm=lm, time_max=time_max)
    # ws = create_lane_vel_sheet(wb=wb, ws=ws, vehlog=vehlog, lm=lm, time_max=time_max, interval=interval)
    create_merging_info_sheet(wb=wb, vehlog=car, lm=lm, )
    ws = create_visual_sheet(wb=wb, ws=ws, vehlog=car, car_info=car_info, lm=lm, time_max=time_max,
                             )  # 可視化
    create_log_sheet(wb=wb, ws=ws, vehlog=car, car_info=car_info, lm=lm, time_max=time_max, )
    # ws = time_vel_sheet(wb, ws, vehlog, car_max, time_max, interval * 10)
    # ws = time_avgvel_sheet(wb, ws, vehlog, car_max, time_max, interval)
    ws = moving_avg_sheet(wb=wb, ws=ws, car=car, lm=lm, time_max=time_max)
    deceleration_log_sheet(wb=wb, car_info=car_info, lm=lm)
    # tracking_log(wb=wb, )
    print(path + " を保存中...")
    print()
    wb.save(path)
"""
